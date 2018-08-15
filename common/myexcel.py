# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os

class MyExcel:
    def __init__(self,filename,logger):
        self.wb = load_workbook(filename)
        self.logger = logger

    #读取变量excel中的变量列表并存入到一个字典中
    def get_init_variables(self,sheetname):
        sh = self.wb.get_sheet_by_name(sheetname)
        init_datas = {}
        for index in range(2,sh.max_row+1):
            # print(index)
            key = sh.cell(index,1).value
            init_datas[key] = str(sh.cell(index,2).value)
        #处理电话号码自增
        # init_datas["${phone}"] = str(int(init_datas["${init_phone}"]) + 1)
        # init_datas["${noReg_phone}"] = str(int(init_datas["${init_phone}"]) + 2)
        # self.logger.info("get_init_variables函数获取到的初始化变量为:{0}".format(init_datas))
        #init_datas:{'${noReg_phone}': '15600000332', '${init_phone}': '15600000330', '${phone}': '15600000331'}
        # print(init_datas)
        return init_datas

    # 读取excel所有行数据，作为ddt参数传入DDT  其中每一行的数据作为ddt传入的一次参数
    #[{'expected_data': '{"status":1,"code":"10001","data":null,"msg":"注册成功"}', 'case_id': '01', 'compare_type': '0', 'request_data': '{"mobilephone":"15600000330","pwd":"0123456789","regname":"zhour"}', '用例说明': '注册成功-有昵称', 'url': 'http://119.23.241.154:8080/futureloan/mvc/api/member/register', 'related_expression': 'None', 'api_name': 'register_success_with_nickname', 'api': '注册', 'http_method': 'get'},{}]
    def getallrow(self,sheetname):
        sh = self.wb.get_sheet_by_name(sheetname)
        one_test_data = []
        dict1 = {}
        cols = sh.max_column
        rows = sh.max_row
        init_datas = self.get_init_variables("variables")
        # print(init_datas)
        for i in range(2,rows+1):
            for j in range(1, cols + 1):
                k = sh.cell(1, j).value
                v = str(sh.cell(i, j).value)
                dict1[k] = v
                for key,value in init_datas.items():
                    # 如果请求数据为空时需要做处理
                    if dict1[k] is not None:
                        ##如果请求数据中包含变量时需要做处理
                        if dict1[k].find(key) != -1:
                            dict1[k] = dict1[k].replace(key,value)
                            #
            one_test_data.append(dict1.copy())
        return one_test_data

    #
    # #更新初始化数据
    # def update_init_data(self,sheetname,i,j):
    #     sh = self.wb.get_sheet_by_name(sheetname)
    #     init_data = self.get_init_variables(sheetname)
    #     sh.cell(i,j).value = str(int(init_data["${init_phone}"]) + 3)
    #     self.logger.info("更新后的初始化数据为：{0}".format(sh.cell(i,j).value))
    #     sh = self.wb.get_sheet_by_name(sheetname)
    #     sh.cell(i, j).value = value
    #     self.logger.info("更新后的动态变量为：{0}".format(sh.cell(i, j).value))
    #
    #     #保存数据
    # def savefile(self,filename):
    #     self.wb.save(filename)